#! python3

import openpyxl, os, re, fnmatch, csv, sys, datetime
from timeit import default_timer as timer
from openpyxl.styles import PatternFill, NamedStyle, Font, Fill

start = timer()

cPath = os.getcwd()
wDir = r'C:\Users\alogan\Desktop\TEMP_GIS\python\tpy\Testing\Compare Testing'

if cPath != wDir:
    os.chdir(wDir)
    cPath = os.getcwd()

pathFolders = os.listdir()
pathFoldersL = len(pathFolders)

u_pathFolders = []

for i in range(0, pathFoldersL):
    if os.path.isdir(pathFolders[i]):
        u_pathFolders.append(pathFolders[i])
        
u_pathFoldersL = len(u_pathFolders)

dir_pathList = []

for i in range(0, u_pathFoldersL):
    newpath = os.path.join(cPath, u_pathFolders[i])
    dir_pathList.append(newpath)

dir_pathListL = len(dir_pathList)

#Paths established
masterPath = dir_pathList[1]
comparePath = dir_pathList[0]
resultsPath = dir_pathList[2]

#Grab Master data
os.chdir(masterPath)

#Find master file
for file in os.listdir('.'):
        if fnmatch.fnmatch(file, '*.xlsx'):
            masterFile = file
masterFile = os.path.join(masterPath, masterFile)

#Find Compare File
os.chdir(comparePath)
for file in os.listdir('.'):
        if fnmatch.fnmatch(file, '*.xlsx'):
            compareFile = file

old_compareFile = file
compareFile = os.path.join(comparePath, compareFile)

print('Opening master file...\n' + str(masterFile))
#Master workbook/sheet
m_wb = openpyxl.load_workbook(masterFile)
m_ws = m_wb.active
print('Opening compare file...\n' + str(compareFile))
#Compare workbook/sheet
c_wb = openpyxl.load_workbook(compareFile, data_only=True)
c_ws = c_wb.active



m_columnHeader = [m_ws.cell(row=1,column=i).value for i in range(1,28)]
#m_columnHeader.append('a')
c_columnHeader = [c_ws.cell(row=1,column=i).value for i in range(1,28)]
#c_columnHeader.append('a')
m_chLen = len(m_columnHeader)
c_chLen = len(c_columnHeader)


"""
#Doesnt appear to be able to find None, delete None after
for i in range(1,28):
    if m_ws.cell(row=1,column=i).value is not 'None':
        m_columnHeader = [m_ws.cell(row=1,column=i).value]
#m_columnHeader.append('a')
for i in range(1,28):
    if c_ws.cell(row=1,column=i).value is not None:
        c_columnHeader = [c_ws.cell(row=1,column=i).value]

m_chLen = len(m_columnHeader)
c_chLen = len(c_columnHeader)

"""

for i in range(0, m_chLen):
    if None in m_columnHeader:
        m_columnHeader.remove(None)
#Remove NoneType
for i in range(0, c_chLen):
    if None in c_columnHeader:
        c_columnHeader.remove(None)





#List of headers to compare
headerList = ['Target_ID', 'Date', 'Team' ,'Ch2_QC_R1', 'Anomaly_Ty', 'Instrument', 'Depth_in',
              'Offset_in', 'Offset_dir', 'Seed_Type', 'Seed_ID', 'Count', 'Weight_lb']


#Change headerList to lower case to avoid case error, search compare must be lowercase too
#THIS LINE CAUSING INDEXERROR, LOOPS BELOW WORK WITHOUT.
headerList = [x.lower() for x in headerList]
m_columnHeader = [y.lower() for y in m_columnHeader]
c_columnHeader = [z.lower() for z in c_columnHeader]

headerListLen = len(headerList)

#Dictionary initialize
m_colH = {}
#c_colH = compare_columnHeader
c_colH = {}

#colNum starts at 1, increment with i
colNum = 1
#Loop through all column headers
for i in range(0, m_chLen):
    #Enter col i -> Loop through list of Header tags
    for k in range(0, headerListLen):
        try:
            if m_columnHeader[i] == headerList[k]:
                #.update acts as append in this scenario
                m_colH.update({m_columnHeader[i] : colNum})
#                print(str(colNum) + ' : ' + str(m_columnHeader[i]))
        except IndexError:
            break
    colNum = colNum + 1
#Reset colNum to 1
colNum = 1

for i in range(0, c_chLen):
    #Enter col i -> Loop through list of Header tags
    for k in range(0, headerListLen):
        try:
#            if c_columnHeader[i] and headerList[k] is not None:
            if c_columnHeader[i] == headerList[k]:
                #.update acts as append in this scenario
                c_colH.update({c_columnHeader[i] : colNum})
#                print(str(colNum) + ' : ' + str(c_columnHeader[i]))
        except IndexError:
            break
    colNum = colNum + 1

#m_colLen = len(m_columnHeader)
#c_colLen = len(c_columnHeader)

#tid = targetid
m_get_tid = m_colH.get('target_id')
c_get_tid = c_colH.get('target_id')

m_get_date = m_colH.get('date')
c_get_date = c_colH.get('date')

#ch2_qc_r1
m_get_ch2 = m_colH.get('ch2_qc_r1')
c_get_ch2 = c_colH.get('ch2_qc_r1')

m_get_anomaly_ty = m_colH.get('anomaly_ty')
c_get_anomaly_ty = c_colH.get('anomaly_ty')

m_get_count = m_colH.get('count')
c_get_count = c_colH.get('count')

m_get_weight_lb = m_colH.get('weight_lb')
c_get_weight_lb = c_colH.get('weight_lb')


#TODO: Implement the following list above m_get_
"""
team
instrument
depth_in
offset_in
offset_dir
seed_type
seed_id
"""









#Stylings, not used
#duplicate_style = PatternFill(fill_type='solid',
#                      start_color='ffc7ce',
#                      end_color='ffc7ce')

#duplicate_style = NamedStyle(name="duplicate")
#c_wb.add_named_style(duplicate_style)


c_rowmax = c_ws.max_row + 1
m_rowmax = m_ws.max_row + 1


date_style = NamedStyle(name='datetime', number_format='DD/MM/YYYY')

print('Formatting dates...')

for i in range(2, c_rowmax):
    c_datecell = c_ws.cell(row=i, column=c_get_date)
    c_datecell_val = c_ws.cell(row=i, column=c_get_date).value
    try:
        c_datecell.style = date_style
    except ValueError:
        pass
    tempdate = c_ws.cell(row=i, column=c_get_date).value
    tempdate = tempdate.replace(tempdate[4], '\\', 1)
    #c_datecell.value = 
"""
#Attempt to copy all values to new file to properly refresh Date Number Format
#Didn't refresh format....
mark_wb = openpyxl.Workbook()
mark_ws = mark_wb.active

for i in range(1, c_rowmax):
    for k in range(1, 28):
        mark_ws.cell(row=i, column=k).value = c_ws.cell(row=i, column=k).value

markupFile = str("MARKING_") + str(old_compareFile)
markupFile = os.path.join(resultsPath, markupFile)

mark_wb.close()
mark_wb.save(markupFile)
"""

#    if c_datecell_val != None:
#        dttm = datetime.datetime.strptime(c_datecell_val, "%Y-%m-%d")
#        c_datecell.value = dttm

#Make Better Ver
"""
for i in range(2, c_rowmax):
    c_datecell = c_ws.cell(row=i, column=c_get_date)
    c_datecell_val = c_ws.cell(row=i, column=c_get_date).value
    if c_datecell != None:
        try:
            c_datecell.style = date_style
            c_ws.cell(row=i, column=c_get_date).value = '0'
#            c_datecell = c_datecell_val
        except ValueError:
            pass
#        c_datecell.number_format = "MM/DD/YY"

        #dttm = datetime.datetime.strptime(c_datecell, "%Y-%m-%d")
        #c_datecell = dttm
#"%Y-%m-%d"    from    "%m/%d/%Y"
    #c_datecell = c_ws.cell(row=i, column=c_get_date)
    #c_datecell.style.number_format = 'MM/DD/YY'
c_wb.close()
c_wb.save(compareFile)
c_wb = openpyxl.load_workbook(compareFile)
c_ws = c_wb.active
"""

#
#TEMP DISABLED ---- WORKS FINE
#

print('Finding duplicates...')
for i in range(2, c_rowmax):
#    print("i : " + str(i))
    for k in range(2, m_rowmax):
#        print("k : " + str(k))
        #print("c : " + str(c_ws.cell(row=i, column=c_get_tid)))
        #print("m : " + str(m_ws.cell(row=k, column=m_get_tid)))
        c_datecell = c_ws.cell(row=i, column=c_get_date)
        c_datecell.number_format = 'MM/DD/YY'
        #m_datecell
        #m_ws.cell(row=k, column=m_get_date).value
        if c_ws.cell(row=i, column=c_get_tid).value == m_ws.cell(row=k, column=m_get_tid).value:
            #Currently just logs:
            #TODO: Implement conditional formatting for positive catch
            #TODO2: if positive result, move on to check date, team, anomaly, etc.
#            print("Row : " + str(i) + " " + str(c_ws.cell(row=i, column=c_get_tid).value) + ' = ' +
#                  "Row : " + str(k) + " " + str(m_ws.cell(row=k, column=m_get_tid).value))
            c_ws.cell(row=i, column=c_get_tid).fill = PatternFill("solid", fgColor="FFC7CE")

            #TODO: NEW LOOP HERE FOR POSITIVE CONTINUATION OF RESULTS
            if c_ws.cell(row=i, column=c_get_date).value != m_ws.cell(row=k, column=m_get_date).value:
                c_ws.cell(row=i, column=c_get_date).fill = PatternFill("solid", fgColor="FFC7CE")


#Works
#test = c_ws['A1']
#test.fill = PatternFill("solid", fgColor="FFC7CE")

end = timer()
print(end - start) 

#old_compareFile
markupFile = str("MARKUP_") + str(old_compareFile)

markupFile = os.path.join(resultsPath, markupFile)

c_wb.close()
c_wb.save(markupFile)

