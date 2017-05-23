#! python3

import openpyxl, os, re, fnmatch, csv, sys, datetime
from timeit import default_timer as timer
from openpyxl.styles import PatternFill, NamedStyle, Font, Fill
from time import sleep

start = timer()

cPath = os.getcwd()
#wDir = r'\\PMP-SRV-INFO009\BO_GSI_Projects\QC Tools\Compare Points'
wDir = r'C:\Users\alogan\Desktop\TEMP_GIS\python\tpy\Testing\Compare_Testing'
if cPath != wDir:
    os.chdir(wDir)
    cPath = os.getcwd()

pathFolders = os.listdir()
pathFoldersL = len(pathFolders)

u_pathFolders = []

for i in range(0, pathFoldersL):
    if os.path.isdir(pathFolders[i]):
        u_pathFolders.append(pathFolders[i])
        
u_pathFoldersL = len(u_pathFolders)

dir_pathList = []

for i in range(0, u_pathFoldersL):
    newpath = os.path.join(cPath, u_pathFolders[i])
    dir_pathList.append(newpath)

dir_pathListL = len(dir_pathList)

#Paths established
masterPath = dir_pathList[1]
comparePath = dir_pathList[0]
resultsPath = dir_pathList[2]

#Grab Master data
os.chdir(masterPath)

#Find master file
for file in os.listdir('.'):
        if fnmatch.fnmatch(file, '*.xlsx'):
            masterFile = file
masterFile = os.path.join(masterPath, masterFile)

#Find Compare File
os.chdir(comparePath)
for file in os.listdir('.'):
        if fnmatch.fnmatch(file, '*.xlsx'):
            compareFile = file

old_compareFile = compareFile
compareFile = os.path.join(comparePath, compareFile)

print('Opening master file...\n' + str(masterFile))
#Master workbook/sheet
m_wb = openpyxl.load_workbook(masterFile)
m_ws = m_wb.active
print('\nOpening compare file...\n' + str(compareFile))
#Compare workbook/sheet
c_wb = openpyxl.load_workbook(compareFile, data_only=True)
c_ws = c_wb.active
print('\nCreating Results Sheet')
#New sheet for markup results
c_wb.create_sheet(index=1, title ='Markup Results')
c_markws = c_wb.get_sheet_by_name('Markup Results')


m_columnHeader = [m_ws.cell(row=1,column=i).value for i in range(1,28)]
#m_columnHeader.append('a')
c_columnHeader = [c_ws.cell(row=1,column=i).value for i in range(1,28)]
#c_columnHeader.append('a')
m_chLen = len(m_columnHeader)
c_chLen = len(c_columnHeader)


"""
#Doesnt appear to be able to find None, delete None after
for i in range(1,28):
    if m_ws.cell(row=1,column=i).value is not 'None':
        m_columnHeader = [m_ws.cell(row=1,column=i).value]
#m_columnHeader.append('a')
for i in range(1,28):
    if c_ws.cell(row=1,column=i).value is not None:
        c_columnHeader = [c_ws.cell(row=1,column=i).value]

m_chLen = len(m_columnHeader)
c_chLen = len(c_columnHeader)

"""

for i in range(0, m_chLen):
    if None in m_columnHeader:
        m_columnHeader.remove(None)


#Remove NoneType
for i in range(0, c_chLen):
    if None in c_columnHeader:
        c_columnHeader.remove(None)

c_chLen = len(c_columnHeader)
for i in range(0, c_chLen):
    if c_columnHeader[i] == 'Anomaly_Type':
        c_columnHeader[i] = 'Anomaly_Ty'
    if c_columnHeader[i] == 'Offset_direction':
        c_columnHeader[i] = 'Offset_Dir'

#List of headers to compare
headerList = ['Target_ID', 'Date', 'Team' ,'Ch2_QC_R1', 'Anomaly_Ty', 'Instrument', 'Depth_in',
              'Offset_in', 'Offset_dir', 'Seed_Type', 'Seed_ID', 'Count', 'Weight_lb']


#Change headerList to lower case to avoid case error, search compare must be lowercase too
#THIS LINE CAUSING INDEXERROR, LOOPS BELOW WORK WITHOUT.
headerList = [x.lower() for x in headerList]
m_columnHeader = [y.lower() for y in m_columnHeader]
c_columnHeader = [z.lower() for z in c_columnHeader]

headerListLen = len(headerList)

#Dictionary initialize
m_colH = {}
#c_colH = compare_columnHeader
c_colH = {}

#colNum starts at 1, increment with i
colNum = 1
#Loop through all column headers
for i in range(0, m_chLen):
    #Enter col i -> Loop through list of Header tags
    for k in range(0, headerListLen):
        try:
            if m_columnHeader[i] == headerList[k]:
                #.update acts as append in this scenario
                m_colH.update({m_columnHeader[i] : colNum})
#                print(str(colNum) + ' : ' + str(m_columnHeader[i]))
        except IndexError:
            break
    colNum = colNum + 1
#Reset colNum to 1
colNum = 1

for i in range(0, c_chLen):
    #Enter col i -> Loop through list of Header tags
    for k in range(0, headerListLen):
        try:
#            if c_columnHeader[i] and headerList[k] is not None:
            if c_columnHeader[i] == headerList[k]:
                #.update acts as append in this scenario
                c_colH.update({c_columnHeader[i] : colNum})
#                print(str(colNum) + ' : ' + str(c_columnHeader[i]))
        except IndexError:
            break
    colNum = colNum + 1

#m_colLen = len(m_columnHeader)
#c_colLen = len(c_columnHeader)

#tid = targetid
m_get_tid = m_colH.get('target_id')
c_get_tid = c_colH.get('target_id')

m_get_date = m_colH.get('date')
c_get_date = c_colH.get('date')

#ch2_qc_r1
m_get_ch2 = m_colH.get('ch2_qc_r1')
c_get_ch2 = c_colH.get('ch2_qc_r1')

m_get_anomaly_ty = m_colH.get('anomaly_ty')
c_get_anomaly_ty = c_colH.get('anomaly_ty')

m_get_count = m_colH.get('count')
c_get_count = c_colH.get('count')

m_get_weight_lb = m_colH.get('weight_lb')
c_get_weight_lb = c_colH.get('weight_lb')

#Additional
m_get_depth_in = m_colH.get('depth_in')
c_get_depth_in = c_colH.get('depth_in')

m_get_offset_in = m_colH.get('offset_in')
c_get_offset_in = c_colH.get('offset_in')

m_get_offset_dir = m_colH.get('offset_dir')
c_get_offset_dir = c_colH.get('offset_dir')

#Markup Worksheet column references - static ref, no need to .get
markcol_tid = 1
markcol_date = 2
markcol_ch2 = 3
markcol_offset_in = 4
markcol_offset_dir = 5
markcol_depth_in = 6
markcol_anomaly_ty = 7 
markcol_weight_lb = 8
markcol_count = 9

#TODO: Implement the following list above m_get_
"""
#team
#instrument
#depth_in
#offset_in
#offset_dir
#seed_type
#seed_id
"""

c_rowmax = c_ws.max_row + 1
m_rowmax = m_ws.max_row + 1


date_style = NamedStyle(name='dateyearmd', number_format='yyyy/mm/dd')
tmpctr = 0
print('Formatting dates...')
for i in range(2, c_rowmax):
#Attempts to transform date from - - to / /
    tempdate = c_ws.cell(row=i, column=c_get_date).value
    try:
        tempdate = tempdate.replace(tempdate[4], '/', 1)
        tempdate = tempdate.replace(tempdate[7], '/', 1)
        c_ws.cell(row=i, column=c_get_date).value = tempdate
        datestrip = datetime.datetime.strptime(tempdate, "%Y/%m/%d")
        c_ws.cell(row=i, column=c_get_date).value = datestrip
    except AttributeError:
        pass
    except TypeError:
        pass


    while tmpctr < 1:
        print('Formatting additional cells...')
        tmpctr = 1
#TODO - Find all No_Find -> change to NO FIND
    try:
        tempanom = c_ws.cell(row=i, column=c_get_anomaly_ty).value
        tempanom = tempanom.upper()
        c_ws.cell(row=i, column=c_get_anomaly_ty).value = tempanom
        if tempanom == 'NO_FIND':
            c_ws.cell(row=i, column=c_get_anomaly_ty).value = 'NO FIND'
    except AttributeError:
        pass
    except TypeError:
        pass
#m_wb.save(masterFile)
#    if c_datecell_val != None:
#        dttm = datetime.datetime.strptime(c_datecell_val, "%Y-%m-%d")
#        c_datecell.value = dttm

#j v for Markup results sheet, row incrementers
v = 2
j = 3
#Set Markup results sheet headers
c_markws.cell(row=1, column=1).value = 'Target_ID'
c_markws.cell(row=1, column=2).value = 'Date'
c_markws.cell(row=1, column=3).value = 'CH2_QC_R1'
c_markws.cell(row=1, column=4).value = 'Offset_in'
c_markws.cell(row=1, column=5).value = 'Offset_dir'
c_markws.cell(row=1, column=6).value = 'Depth_in'
c_markws.cell(row=1, column=7).value = 'Anomaly_Ty'
c_markws.cell(row=1, column=8).value = 'Weight_lb'
c_markws.cell(row=1, column=9).value = 'Count'
#Legend
c_markws.cell(row=1, column=13).value = 'Color'
c_markws.cell(row=1, column=13).fill = PatternFill("solid", fgColor="000000")
c_markws.cell(row=1, column=13).font = Font(b=True, color="FFFFFF")
c_markws.cell(row=2, column=13).value = 'Compare Result'
c_markws.cell(row=3, column=13).value = 'Database'
c_markws.cell(row=3, column=13).fill = PatternFill("solid", fgColor="CCFFCC")
#Format markup header, black fill, bold white text
for i in range(1, 10):
    c_markws.cell(row=1, column=i).fill = PatternFill("solid", fgColor="000000")
    c_markws.cell(row=1, column=i).font = Font(b=True, color="FFFFFF")

#progress bar temp var
cprog = 0
cprog_eq = 0

#Main Loop Starts HERE
print('Sorting duplicates...')
for i in range(2, c_rowmax):
#    print("i : " + str(i))
    for k in range(2, m_rowmax):
#        print("k : " + str(k))
        #print("c : " + str(c_ws.cell(row=i, column=c_get_tid)))
        #print("m : " + str(m_ws.cell(row=k, column=m_get_tid)))
        #Formats Cell dates correctly, removes timestamp
        c_datecell = c_ws.cell(row=i, column=c_get_date)
        c_datecell.number_format = 'MM/DD/YY'
        #m_datecell
        #m_ws.cell(row=k, column=m_get_date).value
        #Compare Target_id
###     TODO: Log both results to markup
        if c_ws.cell(row=i, column=c_get_tid).value == m_ws.cell(row=k, column=m_get_tid).value:
            #Currently just logs:
            #TODO: Implement conditional formatting for positive catch
            #TODO2: if positive result, move on to check date, team, anomaly, etc.
#            print("Row : " + str(i) + " " + str(c_ws.cell(row=i, column=c_get_tid).value) + ' = ' +
#                  "Row : " + str(k) + " " + str(m_ws.cell(row=k, column=m_get_tid).value))
            c_ws.cell(row=i, column=c_get_tid).fill = PatternFill("solid", fgColor="FFEB9C")

            #Markup Results sheet
            c_markws.cell(row=v, column=markcol_tid).value = c_ws.cell(row=i, column=c_get_tid).value
            c_markws.cell(row=j, column=markcol_tid).value = m_ws.cell(row=k, column=m_get_tid).value
            #c_markws.cell(row=v, column=markcol_tid).fill = PatternFill("solid", fgColor="B3B3B3")
            
            try:
                #Date
                if c_ws.cell(row=i, column=c_get_date).value != m_ws.cell(row=k, column=m_get_date).value:
                    c_ws.cell(row=i, column=c_get_date).fill = PatternFill("solid", fgColor="FFC7CE")
                #Markup Results
                    c_markws.cell(row=v, column=markcol_date).value = c_ws.cell(row=i, column=c_get_date).value
                    c_markws.cell(row=j, column=markcol_date).value = m_ws.cell(row=k, column=m_get_date).value
                    #c_markws.cell(row=v, column=markcol_date).fill = PatternFill("solid", fgColor="B3B3B3")
            except TypeError:
                pass

            try:
                #Ch2
                if float(c_ws.cell(row=i, column=c_get_ch2).value) != float(m_ws.cell(row=k, column=m_get_ch2).value):
                    c_ws.cell(row=i, column=c_get_ch2).fill = PatternFill("solid", fgColor="FFC7CE")
                #Markup Results
                    c_markws.cell(row=v, column=markcol_ch2).value = c_ws.cell(row=i, column=c_get_ch2).value
                    c_markws.cell(row=j, column=markcol_ch2).value = m_ws.cell(row=k, column=m_get_ch2).value
                    #c_markws.cell(row=v, column=markcol_ch2).fill = PatternFill("solid", fgColor="B3B3B3")
            except TypeError:
                pass

            try:
                #TODO: Check Anomaly_st?
                #Anomaly_ty
                if c_ws.cell(row=i, column=c_get_anomaly_ty).value != m_ws.cell(row=k, column=m_get_anomaly_ty).value:
                    c_ws.cell(row=i, column=c_get_anomaly_ty).fill = PatternFill("solid", fgColor="FFC7CE")
                #Markup Results
                    c_markws.cell(row=v, column=markcol_anomaly_ty).value = c_ws.cell(row=i, column=c_get_anomaly_ty).value
                    c_markws.cell(row=j, column=markcol_anomaly_ty).value = m_ws.cell(row=k, column=m_get_anomaly_ty).value
                    #c_markws.cell(row=v, column=markcol_anomaly_ty).fill = PatternFill("solid", fgColor="B3B3B3")
            except TypeError:
                pass
            
            try:
                #Count
                if float(c_ws.cell(row=i, column=c_get_count).value) != float(m_ws.cell(row=k, column=m_get_count).value):
                    c_ws.cell(row=i, column=c_get_count).fill = PatternFill("solid", fgColor="FFC7CE")
                #Markup Results
                    c_markws.cell(row=v, column=markcol_count).value = c_ws.cell(row=i, column=c_get_count).value
                    c_markws.cell(row=j, column=markcol_count).value = m_ws.cell(row=k, column=m_get_count).value
                    #c_markws.cell(row=v, column=markcol_count).fill = PatternFill("solid", fgColor="B3B3B3")
            except TypeError:
                pass

            try:
                #Weight_lb
                if float(c_ws.cell(row=i, column=c_get_weight_lb).value) != float(m_ws.cell(row=k, column=m_get_weight_lb).value):
                    c_ws.cell(row=i, column=c_get_weight_lb).fill = PatternFill("solid", fgColor="FFC7CE")
                #Markup Results
                    c_markws.cell(row=v, column=markcol_weight_lb).value = c_ws.cell(row=i, column=c_get_weight_lb).value
                    c_markws.cell(row=j, column=markcol_weight_lb).value = m_ws.cell(row=k, column=m_get_weight_lb).value
                    #c_markws.cell(row=v, column=markcol_weight_lb).fill = PatternFill("solid", fgColor="B3B3B3")
            except TypeError:
                pass

            try:
                #depth_in
                if float(c_ws.cell(row=i, column=c_get_depth_in).value) != float(m_ws.cell(row=k, column=m_get_depth_in).value):
                    c_ws.cell(row=i, column=c_get_depth_in).fill = PatternFill("solid", fgColor="FFC7CE")
                #Markup Results
                    c_markws.cell(row=v, column=markcol_depth_in).value = c_ws.cell(row=i, column=c_get_depth_in).value
                    c_markws.cell(row=j, column=markcol_depth_in).value = m_ws.cell(row=k, column=m_get_depth_in).value
                    #c_markws.cell(row=v, column=markcol_depth_in).fill = PatternFill("solid", fgColor="B3B3B3")
            except TypeError:
                pass
            try:     
                #offset_in
                if float(c_ws.cell(row=i, column=c_get_offset_in).value) != float(m_ws.cell(row=k, column=m_get_offset_in).value):
                    c_ws.cell(row=i, column=c_get_offset_in).fill = PatternFill("solid", fgColor="FFC7CE")
                #Markup Results
                    c_markws.cell(row=v, column=markcol_offset_in).value = c_ws.cell(row=i, column=c_get_offset_in).value
                    c_markws.cell(row=j, column=markcol_offset_in).value = m_ws.cell(row=k, column=m_get_offset_in).value
                    #c_markws.cell(row=v, column=markcol_offset_in).fill = PatternFill("solid", fgColor="B3B3B3")
            except TypeError:
                pass
#            try:       
#                #offset_dir
#                if c_ws.cell(row=i, column=c_get_offset_dir).value != m_ws.cell(row=k, column=m_get_offset_dir).value:
#                    c_ws.cell(row=i, column=c_get_offset_dir).fill = PatternFill("solid", fgColor="FFC7CE")
#                #Markup Results
#                    c_markws.cell(row=v, column=markcol_offset_dir).value = c_ws.cell(row=i, column=c_get_offset_dir).value
#                    c_markws.cell(row=j, column=markcol_offset_dir).value = m_ws.cell(row=k, column=m_get_offset_dir).value
#                    #c_markws.cell(row=v, column=markcol_offset_dir).fill = PatternFill("solid", fgColor="B3B3B3")
#            except TypeError:
#                pass
            
            #increment v odd j even by + 2
            v = v + 2
            j = j + 2

    #Progress Bar
    prog = 100 / float(c_rowmax)
    cprog += prog
    cprogi = int(cprog)
    if cprogi % 5 == 0:
        if cprogi > cprog_eq:
            cprog_eq += cprogi
            cprog_eq = float(cprog_eq) / 5.0
            cprog_eq = int(cprog_eq)
    sys.stdout.write('\r')
    # Progress bar / text output, modify as needed:
    #sys.stdout.write("[%-20s] %d%%" % ('='*i, 5*i))
    sys.stdout.write("[%-23s] %d%%" % ('='*cprog_eq, cprog))
    sys.stdout.flush()

            

print('\nFormatting Markup Results\n')
c_markws_rowmax = c_markws.max_row + 1
# Format Markup Results sheet dates
for i in range(2, c_markws_rowmax):
#Attempts to transform date from - - to / /
    tempdate = c_markws.cell(row=i, column=2).value
    try:
        tempdate = tempdate.replace(tempdate[4], '/', 1)
        tempdate = tempdate.replace(tempdate[7], '/', 1)
        c_markws.cell(row=i, column=markcol_date).value = tempdate
        datestrip = datetime.datetime.strptime(tempdate, "%Y/%m/%d")
        c_markws.cell(row=i, column=markcol_date).value = datestrip
    except AttributeError:
        pass
    except TypeError:
        pass

#Format Markup Dates second pass
for i in range(2, c_markws_rowmax):
    c_markws_datecell = c_markws.cell(row=i, column=2)
    c_markws_datecell.number_format = 'MM/DD/YY'

c_markws_rowmax = float(c_markws_rowmax) / 2.0
c_markws_rowmax = int(c_markws_rowmax) + 2
#Format odd rows as grey fill
for k in range(1, 10):
    j = 3
    for i in range(3, c_markws_rowmax):
        c_markws.cell(row=j, column=k).fill = PatternFill("solid", fgColor="CCFFCC")
        j = j + 2





end = timer()
print("Completed in : " + str(end - start) + "s") 

#old_compareFile
markupFile = str("MARKUP_") + str(old_compareFile)

markupFile = os.path.join(resultsPath, markupFile)

c_wb.close()
c_wb.save(markupFile)

