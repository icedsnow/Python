#! python3

import openpyxl, os, re, fnmatch, csv
from timeit import default_timer as timer

print('Opening workbook...')
currentPath = os.getcwd()
workingDir = 'C:\\Users\\alogan\\Desktop\\TEMP_GIS\\python\\tpy\\Testing\\SSF\\20170426'

if currentPath != workingDir:
    os.chdir(workingDir)
    currentPath = os.getcwd()


gridFile = 'K-8'
gridFilePath = 'C:\\Users\\alogan\\Desktop\\TEMP_GIS\\python\\tpy\\Testing\\SSF\\20170426\\K-8.xlsx'
wb = openpyxl.load_workbook(gridFilePath)
ws = wb.active

print(gridFilePath + '\n' + gridFile + ' is now open...')

def delete_column(ws, delete_column):
    if isinstance(delete_column, str):
        delete_column = openpyxl.cell.column_index_from_string(delete_column)
    assert delete_column >= 1, "Column numbers must be 1 or greater"

    for column in range(delete_column, ws.max_column + 1):
        for row in range(1, ws.max_row + 1):
            ws.cell(row=row, column=column).value = \
                    ws.cell(row=row, column=column+1).value
#Works
def delete_row(ws, delete_row):
    if isinstance(delete_row, str):
        delete_row = openpyxl.cell.row_index_from_string(delete_row)
    assert delete_row >= 1, "Row numbers must be 1 or greater"

    for row in range(delete_row, ws.max_row + 1):
        for column in range(1, ws.max_column + 1):
#            ws.cell(row=row, column=column).value = \
                    ws.cell(row=row+1, column=column).value


def hide_row(ws, row_id):
    if isinstance(row_id, str):
        row_id = openpyxl.cell.row_index_from_string(row_id)
    assert row_id >= 1, "row numbers must be 1 or greater"   
    row_dimension = ws.row_dimensions[row_id]
    row_dimension.hidden = True


newcount = 0
columnHeader = [ws.cell(row=1,column=i).value for i in range(1,28)]
#To line up columnHeader with actual column # do columnHeader + 1
key_list = ['ID', 'Ch2', 'Time', 'Take', 'Photo']
un_key_1 = 'Target_ID'
un_key_2 = 'Seed_ID'
colmax = ws.max_column
#TODO encapsulate in Try Error block (NoneType)
#Try Delete photo columns first?
"""
try:
    delete_column(ws, 25)
    for i in range(20, 28):
        header = ws.cell(row=1,column=i).value
        if 'Field' in header:
            i += 1
        else:
            k = i
            print('PDeleting Column ' + str(k))
            delete_column(ws, k)

except TypeError:
    print('TypeError encountered: Breaking')
"""

#Start of iterator loop to delete columns based on list of predetermined inputs
for i in range(1,5):
    for i in range(1,28):
        header = ws.cell(row=1,column=i).value
        k = i
        try:
            for i in range(0,5):
                    if key_list[i] in header:
                            if un_key_1 not in header and un_key_2 not in header:
                                if header is not None:
    #                                print('Empty column found: ' +str(k))
                                    print('Deleting column ' + str(k))
                                    delete_column(ws, k)
        except TypeError:
                print('TypeError encountered: Breaking')
                break
                        
                        
        columnHeader2 = [ws.cell(row=1,column=i).value for i in range(1,colmax)]

#delete_column(ws, k)

#Find Anomaly_St and CH2_QC_R1
#if Awaiting_Investigation & 9999.99 -> delete row, store in sheet2
colmax = ws.max_column
rowmax = ws.max_row
wb.create_sheet(index=1, title ='Anomaly_QC')
key_anom = 'Awaiting Investigation'
key_ch2 = '9999.99'
#START HERE - FOR LOOP TO ENTER DELETE ROW
cellidList = []
k = 0
start = timer()

"""
#for i in range(1, 10):
for i in range(1, rowmax):
    targetidCell = ws.cell(row=i,column=3).value
    anomCellV = ws.cell(row=i,column=4).value
    ch2CellV = ws.cell(row=i,column=5).value
    if ch2CellV == '9999.999':
        if anomCellV == 'Awaiting_Investigation':
            cellidList.append(targetidCell)
            print('Removing Anomaly: ' + str(cellidList[k]))
            k += 1
            hide_row(ws, i)
print('Finished Hiding Rows -> Now Removing Anomalies')
"""

for i in range(1, 10):
    for i in range(1, rowmax):
        targetidCell = ws.cell(row=i,column=3).value
        anomCellV = ws.cell(row=i,column=4).value
        ch2CellV = ws.cell(row=i,column=5).value
        if ch2CellV == '9999.999':
            if anomCellV == 'Awaiting_Investigation':
                cellidList.append(targetidCell)
                print('Removing Anomaly: ' + str(cellidList[k]))
                k += 1
                delete_row(ws, i)













"""
for i in range(1,colmax):
    header = ws.cell(row=1,column=i).value
    k = i
    try:
        header = ws.cell(row=1,column=i).value

    except TypeError:
                print('TypeError encountered: Breaking')
                break
"""

wb.close()
wb.save('K-8_t.xlsx')
end = timer()
print(end - start) 
#
#
