#! python3

import openpyxl, os, re, fnmatch, csv, sys
from simpledbf import Dbf5
from openpyxl import Workbook

cPath = os.getcwd()
wDir = r'C:\Users\alogan\Desktop\TEMP_GIS\python\tpy\Testing\SSF'

#Set cPath ONLY with os.getcwd()
if cPath != wDir:
    os.chdir(wDir)
    cPath = os.getcwd()

pathFiles = os.listdir()
pathFilesL = len(pathFiles)

pathDateRegex = re.compile(r'\'\d\d\d\d\d\d\d\d\'')

u_pathFiles = []

for i in range(0, pathFilesL):
    if os.path.isdir(pathFiles[i]):
        u_pathFiles.append(pathFiles[i])

pathFilesFind = pathDateRegex.findall(str(pathFiles))
pathFilesFindLen = len(pathFilesFind)
u_pathFiles = []

for i in range(0, pathFilesFindLen):
    folderStrip = pathFilesFind[i]
    folderStrip = folderStrip.strip('\'')
    u_pathFiles.append(folderStrip)

u_pathFilesL = len(u_pathFiles)

#dirPathList = All selected full paths
dir_pathList = []

#.join more efficient way of handling paths
for i in range(0, u_pathFilesL):
    newpath = os.path.join(cPath, u_pathFiles[i])
    dir_pathList.append(newpath)

dir_pathListL = len(dir_pathList)



#iterator wb/ws
y = 1
wb_i = []
ws_i = []
wb_i.append('0')
ws_i.append('0')






#Start loop, ending at the last entry in dir_pathList
for i in range(0, dir_pathListL):
    
#1 - Change directory to first path in list
    os.chdir(dir_pathList[i])

#2 - Update our current path link
    cPath = os.getcwd()
    
#3 - Files found -> enter new loop  to convert current directory from .dbf to .csv    
    all_files = os.listdir('.')
    dbfs = [f for f in all_files if f.endswith('dbf')]
    for dbf in dbfs:
        data = Dbf5(dbf)
        data.to_csv(dbf[:-3] + 'csv')

#4 - Prep openpyxl workbook and sheet
    wb = openpyxl.Workbook()
    ws = wb.active

#5 - Files found -> enter new loop to convert current directory from .csv to .xlsx                
    for file in os.listdir('.'):
        if fnmatch.fnmatch(file, '*.csv'):

            #Testing of iterator on wb/ws
            wb_i.append(Workbook())
            newsheet = wb_i[y].create_sheet()
            ws_i.append(newsheet)
            ws_i[y] = wb_i[y].worksheets[0]
            ws_current = ws_i[y]

            csvStrip = file.strip('.csv')
            csvPath = os.path.join(cPath, file)
            xlsxPath = str(csvStrip) + '.xlsx'
            #csvFile = open(file)
            with open(file) as csvFile:
                csvReader = csv.reader(csvFile, delimiter=',')
                for row in csvReader:
                    ws_current.append(row)

            #del can probably be removed, with(open) as f should do this
            #del temp ensuring files are gone
            del csvReader
            
            csvFile.close()
            del csvFile
            del file
            
            wb_current = wb_i[y]
            wb_current.close()
            wb_current.save(xlsxPath)
            y = y + 1
            #wb.save(xlsxPath)

#End of Line
print('EOL')










