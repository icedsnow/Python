#! python3

import openpyxl, os, re, fnmatch, csv
from openpyxl import Workbook
from timeit import default_timer as timer
#import multiprocessing

print('Opening workbook...')
cPath = os.getcwd()
wDir = 'C:\\Users\\alogan\\Desktop\\TEMP_GIS\\python\\tpy\\Testing\\SSF'

if cPath != wDir:
    os.chdir(wDir)
    cPath = os.getcwd()

def delete_column(ws, delete_column):
    if isinstance(delete_column, str):
        delete_column = openpyxl.cell.column_index_from_string(delete_column)
    assert delete_column >= 1, "Column numbers must be 1 or greater"

    for column in range(delete_column, ws.max_column + 1):
        for row in range(1, ws.max_row + 1):
            ws.cell(row=row, column=column).value = \
                    ws.cell(row=row, column=column+1).value
#Works
def delete_row(ws, delete_row):
    if isinstance(delete_row, str):
        delete_row = openpyxl.cell.row_index_from_string(delete_row)
    assert delete_row >= 1, "Row numbers must be 1 or greater"

    for row in range(delete_row, ws.max_row + 1):
        for column in range(1, ws.max_column + 1):
            ws.cell(row=row, column=column).value = \
                    ws.cell(row=row+1, column=column).value

#TODO open files dynamically

pathFiles = os.listdir()
pathFilesL = len(pathFiles)
    
pathDateRegex = re.compile(r'\'\d\d\d\d\d\d\d\d\'')

u_pathFiles = []

for i in range(0, pathFilesL):
    if os.path.isdir(pathFiles[i]):
        u_pathFiles.append(pathFiles[i])

pathFilesFind = pathDateRegex.findall(str(pathFiles))
pathFilesFindLen = len(pathFilesFind)
u_pathFiles = []

for i in range(0, pathFilesFindLen):
    folderStrip = pathFilesFind[i]
    folderStrip = folderStrip.strip('\'')
    u_pathFiles.append(folderStrip)

u_pathFilesL = len(u_pathFiles)

#dirPathList = All selected full paths
dir_pathList = []


#.join more efficient way of handling paths
for i in range(0, u_pathFilesL):
    newpath = os.path.join(cPath, u_pathFiles[i])
    dir_pathList.append(newpath)

dir_pathListL = len(dir_pathList)
#y iterator for ws_i[y]
#Start loop, ending at the last entry in dir_pathList
for z in range(0, dir_pathListL):
    
    #1 - Change directory to first path in list
    os.chdir(dir_pathList[z])

    #2 - Update our current path link
    cPath = os.getcwd()

    #3 - Prep openpyxl workbook and sheet
    wb = openpyxl.Workbook()
    ws = wb.active

    #4 - Files found -> enter new loop to find xlsx files
    for file in os.listdir('.'):
        if fnmatch.fnmatch(file, '*.xlsx'):
            xlsxStrip = file.strip('.xlsx')
            joinPath = os.path.join(cPath, file)
            xlsxPath = str(xlsxStrip) + '.xlsx'
            xlsxFile = open(file)

            gridFile = 'xlsxFile'
            gridFilePath = xlsxPath
            wb = openpyxl.load_workbook(gridFilePath)
            ws = wb.active

            print(gridFilePath + '\n' + gridFile + ' is now open...')



            newcount = 0
            columnHeader = [ws.cell(row=1,column=i).value for i in range(1,28)]
            key_list = ['ID', 'Ch2', 'Time', 'Take', 'Photo']
            un_key_1 = 'Target_ID'
            un_key_2 = 'Seed_ID'
            colmax = ws.max_column

            #Start of iterator loop to delete columns based on list of predetermined inputs
            for i in range(1,5):
                for i in range(1,28):
                    header = ws.cell(row=1,column=i).value
                    k = i
                    try:
                        for i in range(0,5):
                                if key_list[i] in header:
                                        if un_key_1 not in header and un_key_2 not in header:
                                            if header is not None:
                #                                print('Empty column found: ' +str(k))
                                                print('Deleting column ' + str(k))
                                                delete_column(ws, k)
                    except TypeError:
                            print('TypeError encountered: Breaking')
                            break
                                    
                                    
                    columnHeader2 = [ws.cell(row=1,column=i).value for i in range(1,colmax)]


            colmax = ws.max_column
            rowmax = ws.max_row
            wb.create_sheet(index=1, title ='Anomaly_QC')

            #START HERE - FOR LOOP TO ENTER DELETE ROW
            cellidList = []
            k = 0
            #Start Timer
            start = timer()

#wb_i defined here
            y = z + 1
            wb_i = []
            ws_i = []
            wb_i.append('0')
            ws_i.append('0')
            wb_i.append(Workbook())
            newsheet = wb_i[y].create_sheet()
            ws_i.append(newsheet)
            ws_i[y] = wb_i[y].worksheets[0]



            """TESTING wb iterator use
                        wb2 = Workbook()
                        wb2.create_sheet()    #(index=0, title ='Master')
                        ws2 = wb2.worksheets[0]
                        #wb2.save('')
            """


            #k = column value From ws1
            k = 1
            row_nocopy = 0
            i2 = 0
            #Enter loop, Rows 1 -> max row + 1
            for i in range(1, rowmax + 1):
                #Enter 2nd loop, iterate across columns, moving down rows once all cols done
                for k in range(1, 28):
                    targetidCell = ws.cell(row=i,column=3).value
                    anomCellV = ws.cell(row=i,column=4).value
                    ch2CellV = ws.cell(row=i,column=5).value
                    ws_current = ws_i[y]
                    ws_current.cell(row=1, column=k).value = \
                        ws.cell(row=1, column=k).value
                    
                    if ch2CellV < '9999' and anomCellV != 'Awaiting_Investigation':
                        ws_current.cell(row=i2, column=k).value = \
                            ws.cell(row=i, column=k).value

                    #Else should only run if no copy occures, keeping i2 in sync with i
                if ch2CellV > '9999' and anomCellV == 'Awaiting_Investigation':
                    row_nocopy +=1
                i2 += 1
                i2 -= row_nocopy
                row_nocopy = 0

                k = 1 #1 indent, attached to rowmax loop
            rowmax2 = ws_current.max_row
            #Run delete to clear blank white spaces missed
            k = 0
            for i in range(1, 10):
                for i in range(1, rowmax2 + 1):
                    targetidCell = ws_current.cell(row=i,column=3).value
                    if targetidCell is None:
                        cellidList.append(targetidCell)
                        print('Removing Anomaly: #' + str(k) + ' = i:' + str(i))
                        k += 1
                        delete_row(ws_current, i)

            wb.close()
            
            wb_current = wb_i[y]
            wb_current.close()
            wb_current.save(xlsxPath)
end = timer()
print(end - start) 


"""
#Breakpoint testing for variables above this line
while True:
    BREAK = input("RUNTIME STOPPED FROM SOURCE")
"""                              
